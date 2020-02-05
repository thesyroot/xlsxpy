from openpyxl import Workbook, load_workbook

class xlsxpy:

    def __init__(self, namex="HojaCalculo.xlsx"):
        try:
            self.__file = load_workbook(namex)
        except FileNotFoundError:
            self.__file = Workbook();
            self.__file.save(filename = namex)
            self.__file = load_workbook(namex)
        self.__sheet = self.__file.active
        self.__namef = namex

    def Read(self, celda="-1"):
        try:
            return str(self.__sheet[celda].value)
        except AttributeError:
            print("Maybe you write wrong")
            return False

    def Write(self, text, celda="-1"):
        if(celda == "-1"):
            self.__contenido = []
            # Interpretando texto
            for elemento in text.split("\n"):
                self.__contenido.append(elemento.split("#/c"))
            # Ingresandolo
            for row in self.__contenido:
                self.__sheet.append(row)
        else:
            try:
                self.__sheet[celda] = text
            except AttributeError:
                print("Maybe you write wrong")

    def cSheet(self, Iden):
        if(Iden.isdigit()):
            if(Iden > 0):
                self.__sheet = self.__file.get_sheet_by_name(
                    self.__file.sheetnames[Iden-1])
                return True
            else:
                print("The id should be greater than 0")
                return False
        else:
            try:
                self.__sheet = self.__file.get_sheet_by_name(Iden)
                return True
            except KeyError:
                print("Maybe you write wrong")
                return False

    def mkSheet(self, Iden):
        self.__file.create_sheet(Iden)

    def rmSheet(self, Iden):
        if(Iden.isdigit()):
            if(Iden > 0):
                self.__file.remove(self.__file.get_sheet_by_name(
                    self.__file.sheetnames[Iden-1]))
                return True
            else:
                print("The id should be greater than 0")
                return False
        else:
            try:
                self.__file.remove(self.__file.get_sheet_by_name(Iden))
                return True
            except KeyError:
                print("Maybe you write wrong")
                return False

    def getSheet(self):
        return str(self.__sheet).split("\"")[1]

    def save(self):
        self.__file.save(self.__namef)
